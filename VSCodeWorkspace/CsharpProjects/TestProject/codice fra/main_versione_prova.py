from os import name
import networkx as nx
from matplotlib import pyplot as plt
import math
import numpy as np
import random
import copy
import statistics as stat
from itertools import islice
import csv
from openpyxl import Workbook


def k_shortest_paths(G, source, target, k, weight=None):
    try:
        paths = islice(nx.shortest_simple_paths(G, source, target, weight=weight), k)
        return list(paths)
    except nx.exception.NetworkXNoPath:
        return False

def fidelity_best_path(G, source, target, virtual_links):

  def find_virtual_link(a, b, virtual_links):
    for virtual_link in virtual_links:
        if virtual_link.active_link == (a, b) or virtual_link.active_link == (b, a):
            return virtual_link
    return None

  max_fidelity = 0
  best_path = []
  for path in nx.all_simple_paths(G, source, target):
    print(path)
    sum_fidelity_path = 0
    for i in range(0,len(path)-1,1):
      current_i = path[i]
      after_i = path[i+1]

      current_virtual_link = find_virtual_link(current_i, after_i, virtual_links)
      print(current_virtual_link.fidelity_sublinks[0])

      sum_fidelity_path += current_virtual_link.fidelity_sublinks[0]

    if sum_fidelity_path >= max_fidelity:
      max_fidelity = sum_fidelity_path
      best_path = path

  if best_path:
    print(f"Max fidelity: {max_fidelity}")
    return [best_path]
  else:
    return False

def all_paths(G, source, target, cutoff=None):
        return list(nx.all_simple_paths(G, source, target, cutoff))

# def purification(x1, x2, purification_probability):
#        if np.random.binomial(1, purification_probability) == 1:
#         return x1 * x2 / (x1 * x2 + (1 - x1) * (1 - x2))
#        else:
#         return -1 #Non c'è purificazione

def purification(x1, x2, pur_round, purification_probability):
       initial_fidelity = 0.6
       for i in range(1, pur_round+1):
        if np.random.binomial(1, purification_probability) == 1:
            return x1 * x2 / (x1 * x2 + (1 - x1) * (1 - x2))
        else:
            return -1 #Non c'è purificazione

class VirtualLink:
    def __init__(self, a, b, timer, fidelity_sublinks, old_physical_links = []) -> None:
        self.active_link = (a,b)
        self.timer = timer
        self.fidelity_sublinks = fidelity_sublinks
        self.old_physical_links = old_physical_links

    def get_active_link(self):
        return self.active_link

    def get_old_physical_link(self):
        return self.old_physical_link

    def get_timer(self):
        return self.timer

    def set_timer(self, t):
        self.timer = t

    def decrease_timer(self, t):
        self.timer = self.timer - t


class Graph:
    def __init__(self, nodes, physical_links, p_gen, p_swap, t_dec, t_gen, t_swap) -> None:
        self.nodes = nodes
        self.physical_links = physical_links
        self.virtual_links = []
        self.p_gen = p_gen
        self.p_swap = p_swap
        self.t_dec = t_dec
        self.t_gen = t_gen
        self.t_swap = t_swap

    def get_path_virtual_links(path):
            path_virtual_links = []
            for i in range(0,len(path)-1):
                path_virtual_links.append((path[i], path[i+1]))
            return path_virtual_links

    def find_virtual_link(self, a, b):
        for virtual_link in self.virtual_links:
            if virtual_link.active_link == (a, b) or virtual_link.active_link == (b, a):
                return virtual_link
        return None

    def entanglement_generation_procedure(self, purification_probability):
        for physical_link in self.physical_links:
            link_exists = False
            for virtual_link in self.virtual_links:
                if physical_link == virtual_link.active_link or physical_link in virtual_link.old_physical_links:
                    link_exists = True
                    if virtual_link.timer - self.t_gen > 0:
                        virtual_link.decrease_timer(self.t_gen)
                    else:
                        self.virtual_links.remove(virtual_link)
                    break
            if not link_exists and np.random.binomial(1, self.p_gen) == 1:
                fidelity12 = purification(0.6, 0.6, purification_probability)

                if fidelity12 == -1:
                    self.virtual_links.append(VirtualLink(physical_link[0], physical_link[1], self.t_dec, [0.6]))
                else:
                    fidelity_tot = purification(fidelity12, 0.6, purification_probability)
                    if fidelity_tot != -1:
                        self.virtual_links.append(VirtualLink(physical_link[0], physical_link[1], self.t_dec, [fidelity_tot]))

    def swapping_procedure(self, path, iter) -> bool:
        succ = True
        path_temp = path.copy()
        while succ and len(path) > 2:
            if len(path) % 2 == 0:
                for i in range(1,len(path)-1,2):
                    current_i = path[i]
                    previous_i = path[i-1]
                    after_i = path[i+1]

                    prev_current_virtual_link = self.find_virtual_link(previous_i, current_i)
                    curr_after_virtual_link = self.find_virtual_link(current_i, after_i)

                    if prev_current_virtual_link.timer  > self.t_swap and curr_after_virtual_link.timer > self.t_swap:
                        if np.random.binomial(1, self.p_swap) == 1:

                            old_links = prev_current_virtual_link.old_physical_links + curr_after_virtual_link.old_physical_links \
                                        + [(previous_i, current_i), (current_i, after_i)]
                            new_timer = min(prev_current_virtual_link.timer, curr_after_virtual_link.timer)
                            new_fidelity = min(prev_current_virtual_link.fidelity_sublinks, curr_after_virtual_link.fidelity_sublinks)

                            self.virtual_links.append(VirtualLink(previous_i, after_i, new_timer - self.t_swap, new_fidelity, old_links))
                            self.virtual_links.remove(prev_current_virtual_link)
                            self.virtual_links.remove(curr_after_virtual_link)
                            path_temp.remove(current_i)
                        else:
                            self.virtual_links.remove(prev_current_virtual_link)
                            self.virtual_links.remove(curr_after_virtual_link)
                            succ = False
                    elif prev_current_virtual_link.timer < self.t_swap:
                        self.virtual_links.remove(prev_current_virtual_link)
                        succ = False
                    elif curr_after_virtual_link.timer < self.t_swap:
                        self.virtual_links.remove(curr_after_virtual_link)
                        succ = False

                iter += 1
                path = path_temp.copy()
                final_link = self.find_virtual_link(path[-1], path[-2])
                final_link.decrease_timer(t_swap)
            else:
                for i in range(1,len(path),2):
                    current_i = path[i]
                    previous_i = path[i-1]
                    after_i = path[i+1]

                    prev_current_virtual_link = self.find_virtual_link(previous_i, current_i)
                    curr_after_virtual_link = self.find_virtual_link(current_i, after_i)

                    if prev_current_virtual_link.timer  > self.t_swap and curr_after_virtual_link.timer > self.t_swap:
                        if np.random.binomial(1, self.p_swap) == 1:

                            old_links = prev_current_virtual_link.old_physical_links + curr_after_virtual_link.old_physical_links \
                                        + [(previous_i, current_i), (current_i, after_i)]
                            new_timer = min(prev_current_virtual_link.timer, curr_after_virtual_link.timer)
                            new_fidelity = min(prev_current_virtual_link.fidelity_sublinks, curr_after_virtual_link.fidelity_sublinks)

                            self.virtual_links.append(VirtualLink(previous_i, after_i, new_timer - self.t_swap, new_fidelity, old_links))
                            self.virtual_links.remove(prev_current_virtual_link)
                            self.virtual_links.remove(curr_after_virtual_link)
                            path_temp.remove(current_i)
                        else:
                            self.virtual_links.remove(prev_current_virtual_link)
                            self.virtual_links.remove(curr_after_virtual_link)
                            succ = False
                    elif prev_current_virtual_link.timer < self.t_swap:
                        self.virtual_links.remove(prev_current_virtual_link)
                        succ = False
                    elif curr_after_virtual_link.timer < self.t_swap:
                        self.virtual_links.remove(curr_after_virtual_link)
                        succ = False
                iter += 1
                print(f"Path: {path}")
                print(f"Path_temp: {path_temp}")
                path = path_temp.copy()

        return path, iter

    def establish_entanglement(self, source, target, purification_probability):
        G = nx.Graph()
        G.add_nodes_from(self.nodes)

        iter_generation = 0
        iter_swapping = 0
        while True:
            path = k_shortest_paths(G, source, target, 1)
            # path = fidelity_best_path(G, source, target, self.virtual_links)
            print(path)

            if path is False:
                self.entanglement_generation_procedure(purification_probability)
                virtual_links = [vl.active_link for vl in self.virtual_links]
                print(virtual_links)
                G.add_edges_from(virtual_links)

                iter_generation += 1
            else:
                print(path[0])
                self.decrease_swapping_time(path[0])
                new_path, it = self.swapping_procedure(path[0], iter_swapping)
                iter_swapping += it

                if len(new_path) == 2 and new_path[0] == SOURCE and  new_path[1] == TARGET:
                    print("End to end entanglement established")
                    print(f"Decoherence time {self.find_virtual_link(source, target).timer}")
                    break
                else:
                    G.remove_edges_from(list(G.edges))
                    virtual_links = [vl.active_link for vl in self.virtual_links]
                    G.add_edges_from(virtual_links)

        return self.t_gen*iter_generation, self.t_swap*iter_swapping

    def decrease_swapping_time(self, path):

        def get_path_virtual_links(path):
            path_virtual_links = []
            for i in range(0,len(path)-1):
                path_virtual_links.append((path[i], path[i+1]))
            return path_virtual_links

        path_links = get_path_virtual_links(path)

        for virtual_link in self.virtual_links:
            if virtual_link.active_link not in path_links:
                if virtual_link.timer > self.t_swap:
                    virtual_link.decrease_timer(self.t_swap)
                else:
                    self.virtual_links.remove(virtual_link)

    def get_all_virtual_links(self):
        return [x.get_active_link() for x in self.virtual_links]
    

def plot_variando_le_probabilità(iter):
    results = []
    for p_gen in [0.5, 0.6, 0.7, 0.8, 0.9, 1]:
        for p_swap in [0.5, 0.6, 0.7, 0.8, 0.9, 1]:
            for pur_prob in [0.5, 0.6, 0.7, 0.8, 0.9, 1]:
                fidelity_array = np.zeros(iter)
                for i in range(iter):
                    g = Graph(nodes, physical_links, p_gen, p_swap, t_dec, t_gen, t_swap)
                    tempo_gen, tempo_swapping = g.establish_entanglement(SOURCE, TARGET, pur_prob)
                    fidelity = g.find_virtual_link(SOURCE, TARGET).fidelity_sublinks
                    fidelity_array[i] = fidelity[0]
                    del g
                results.append([p_gen, p_swap, pur_prob, fidelity_array.mean()])

    csv_file = 'results.csv'
    with open(csv_file, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(['Probabilità di purificazione', 'Fidelity'])
        writer.writerows(results)

    wb = Workbook()
    ws = wb.active
    with open(csv_file, 'r') as f:
        for row in csv.reader(f):
            ws.append(row)

    excel_file = 'results.xlsx'
    wb.save(excel_file)
    print(f'Risultati salvati con successo in {excel_file} e {csv_file}')


def plot_variando_solo_la_probabilità_di_purificazione(nodes, physical_links, p, p_swap, t_dec, t_gen, t_swap, iter):
    results = []
    for pur_prob in [0.5, 0.6, 0.7, 0.8, 0.9, 1]:
        fidelity_array = np.zeros(iter)
        for i in range(iter):
            g = Graph(nodes, physical_links, p, p_swap, t_dec, t_gen, t_swap)
            tempo_gen, tempo_swapping = g.establish_entanglement(SOURCE, TARGET, pur_prob)
            fidelity = g.find_virtual_link(SOURCE, TARGET).fidelity_sublinks
            fidelity_array[i] = fidelity[0]
            del g
        results.append([p_gen, p_swap, pur_prob, fidelity_array.mean()])

    csv_file = 'results.csv'
    with open(csv_file, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(['Probabilità di purificazione', 'Fidelity'])
        writer.writerows(results)

    wb = Workbook()
    ws = wb.active
    with open(csv_file, 'r') as f:
        for row in csv.reader(f):
            ws.append(row)

    excel_file = 'results.xlsx'
    wb.save(excel_file)
    print(f'Risultati salvati con successo in {excel_file} e {csv_file}')

def plot_times(nodes, physical_links, p_prop, p_swap, t_dec, t_gen, t_swap, iter):
    results = []
    for p_gen in [0.5, 0.6, 0.7, 0.8, 0.9, 1]:
        p = p_gen*p_prop
        for p_swap in [0.5, 0.6, 0.7, 0.8, 0.9, 1]:
            for pur_prob in [0.5, 0.6, 0.7, 0.8, 0.9, 1]:
                time_gen = np.zeros(iter)
                time_swap = np.zeros(iter)
                for i in range(iter):
                    g = Graph(nodes, physical_links, p, p_swap, t_dec, t_gen, t_swap)
                    tempo_gen, tempo_swapping = g.establish_entanglement(SOURCE, TARGET, pur_prob)
                    time_gen[i] = tempo_gen
                    time_swap[i] = tempo_swapping
                    del g
                results.append([p_gen, p_swap, pur_prob, time_gen.mean(), time_swap.mean()])

    csv_file = 'time/results_times.csv'
    with open(csv_file, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(['Probabilità di Generazione', 'Probabilità di Swapping', 'Probabilità di purificazione', 'Tempo di generazione', 'Tempo di swapping'])
        writer.writerows(results)

    wb = Workbook()
    ws = wb.active
    with open(csv_file, 'r') as f:
        for row in csv.reader(f):
            ws.append(row)

    excel_file = 'time/results_time.xlsx'
    wb.save(excel_file)
    print(f'Risultati salvati con successo in {excel_file} e {csv_file}')



if __name__ == "__main__":

    nodes = [1,2,3,4,5]
    physical_links = [(1,2), (2,3), (3,4), (4,5), (5,1), (2,4), (2,5)]

    t_dec = 200e-3

    d = 1 #distanza in km tra la sorgente e i nodi

    p_gen = 1; #probabilità di successo nella generazione dell'entanglement (tecnologia)

    t_prop = d/(2e5); #tempo richiesto per il passaggio del fotone nella fibra ottica [s]
    t_gen = 0.1e-6+1/2*t_prop; #tempo richiesto per un tentativo di generazione dell'entanglement [s] (tecnologia)
    print(t_gen)

    p_init = 0.2
    ni = 0.2 #fattore di attenuazione lungo la fibra ottica [dB/km]
    p_prop = (1-p_init)*10**(-ni*d/10) #probabilità che il fotone sia inviato con successo preso dal paper di cicconetti
    
    # p = p_gen*p_prop #probabilità congiunta che tiene conto della p_gen e p_prop

    p_swap = 1 #probabilità di successo di un'operazione di swapping o il contrario???
    t_swap = 100e-6

    SOURCE = 1
    TARGET = 5

    iter = 500

    # plot_variando_le_probabilità(nodes, physical_links, p, p_swap, t_dec, t_gen, t_swap, iter)

    plot_times(nodes, physical_links, p_prop, p_swap, t_dec, t_gen, t_swap, iter)